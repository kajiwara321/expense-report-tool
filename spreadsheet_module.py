"""
経費レポート自動化ツール - スプレッドシートモジュール

このモジュールは、経費データをExcelスプレッドシートに
出力する機能を提供します。
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging

# ロギングの設定
logging.basicConfig(
    level=logging.INFO if os.getenv("DEBUG", "false").lower() == "true" else logging.WARNING,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("spreadsheet_module")

class ExpenseSpreadsheet:
    """経費データをExcelスプレッドシートに出力するクラス"""
    
    def __init__(self, output_path=None):
        """
        スプレッドシートモジュールの初期化
        
        Args:
            output_path (str, optional): 出力ファイルパス。指定しない場合は自動生成。
        """
        if output_path is None:
            # デフォルトのファイル名（現在の日時を含む）
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"expense_report_{current_time}.xlsx"
        
        self.output_path = output_path
        logger.info(f"ExpenseSpreadsheet initialized with output path: {output_path}")
    
    def create_new_workbook(self):
        """
        新しいワークブックを作成
        
        Returns:
            openpyxl.Workbook: 作成されたワークブック
        """
        try:
            workbook = openpyxl.Workbook()
            # デフォルトシートの名前を変更
            sheet = workbook.active
            sheet.title = "経費レポート"
            logger.info("New workbook created")
            return workbook
        except Exception as e:
            logger.error(f"Error creating new workbook: {str(e)}")
            raise
    
    def load_workbook(self, file_path):
        """
        既存のワークブックを読み込み
        
        Args:
            file_path (str): 読み込むファイルのパス
            
        Returns:
            openpyxl.Workbook: 読み込まれたワークブック
        """
        try:
            if os.path.exists(file_path):
                workbook = openpyxl.load_workbook(file_path)
                logger.info(f"Workbook loaded from {file_path}")
                return workbook
            else:
                logger.warning(f"File {file_path} does not exist, creating new workbook")
                return self.create_new_workbook()
        except Exception as e:
            logger.error(f"Error loading workbook: {str(e)}")
            raise
    
    def apply_header_style(self, sheet, row, start_col, end_col):
        """
        ヘッダー行にスタイルを適用
        
        Args:
            sheet: ワークシート
            row (int): ヘッダー行の番号
            start_col (int): 開始列
            end_col (int): 終了列
        """
        # ヘッダーのスタイル
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 罫線スタイル
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダー行にスタイルを適用
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def apply_data_style(self, sheet, start_row, end_row, start_col, end_col):
        """
        データ行にスタイルを適用
        
        Args:
            sheet: ワークシート
            start_row (int): 開始行
            end_row (int): 終了行
            start_col (int): 開始列
            end_col (int): 終了列
        """
        # データ行のスタイル
        data_alignment = Alignment(vertical="center")
        
        # 罫線スタイル
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 交互の行の背景色
        even_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # データ行にスタイルを適用
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # 偶数行に背景色を適用
                if row % 2 == 0:
                    cell.fill = even_fill
    
    def apply_total_style(self, sheet, row, col):
        """
        合計行にスタイルを適用
        
        Args:
            sheet: ワークシート
            row (int): 合計行の番号
            col (int): 合計列の番号
        """
        # 合計のスタイル
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        total_alignment = Alignment(horizontal="right", vertical="center")
        
        # 罫線スタイル
        medium_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        # 合計セルにスタイルを適用
        cell = sheet.cell(row=row, column=col)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = total_alignment
        cell.border = medium_border
    
    def adjust_column_width(self, sheet):
        """
        列幅を内容に合わせて調整
        
        Args:
            sheet: ワークシート
        """
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # 列番号を取得
            
            for cell in col:
                if cell.value:
                    # セルの値の長さを計算（日本語は2文字分としてカウント）
                    try:
                        length = len(str(cell.value))
                        # 日本語文字のカウント（簡易的な方法）
                        for char in str(cell.value):
                            if ord(char) > 127:  # ASCII範囲外の文字
                                length += 1
                        max_length = max(max_length, length)
                    except:
                        pass
            
            # 最小幅を8文字、最大幅を50文字に設定
            adjusted_width = max(8, min(50, max_length + 2))
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width
    
    def write_expense_data(self, expense_data_list, output_path=None):
        """
        経費データをスプレッドシートに書き込み
        
        Args:
            expense_data_list (list): 経費データのリスト
            output_path (str, optional): 出力ファイルパス。指定しない場合はインスタンス作成時のパスを使用。
            
        Returns:
            str: 保存されたファイルのパス
        """
        if output_path is not None:
            self.output_path = output_path
        
        try:
            # 新しいワークブックを作成
            workbook = self.create_new_workbook()
            sheet = workbook.active
            
            # ヘッダー行の作成
            headers = ["No.", "日付", "説明", "金額", "備考"]
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col).value = header
            
            # ヘッダーにスタイルを適用
            self.apply_header_style(sheet, 1, 1, len(headers))
            
            # データの書き込み
            total_amount = 0
            for i, expense in enumerate(expense_data_list, 1):
                row = i + 1  # ヘッダー行の後から開始
                
                # No.
                sheet.cell(row=row, column=1).value = i
                
                # 日付
                sheet.cell(row=row, column=2).value = expense.get("date", "")
                
                # 説明
                sheet.cell(row=row, column=3).value = expense.get("description", "")
                
                # 金額
                amount = expense.get("amount", 0)
                sheet.cell(row=row, column=4).value = amount
                # 数値フォーマットを適用（通貨形式）
                sheet.cell(row=row, column=4).number_format = "¥#,##0"
                
                # 備考（カテゴリなど）
                sheet.cell(row=row, column=5).value = expense.get("category", "")
                
                # 合計金額の計算
                total_amount += amount
            
            # データ行にスタイルを適用
            if expense_data_list:
                self.apply_data_style(sheet, 2, len(expense_data_list) + 1, 1, len(headers))
            
            # 合計行の追加
            total_row = len(expense_data_list) + 3
            sheet.cell(row=total_row, column=3).value = "合計"
            sheet.cell(row=total_row, column=4).value = total_amount
            sheet.cell(row=total_row, column=4).number_format = "¥#,##0"
            
            # 合計行にスタイルを適用
            self.apply_total_style(sheet, total_row, 4)
            
            # 列幅の調整
            self.adjust_column_width(sheet)
            
            # ファイルの保存
            workbook.save(self.output_path)
            logger.info(f"Expense data written to {self.output_path}")
            
            return self.output_path
        except Exception as e:
            logger.error(f"Error writing expense data: {str(e)}")
            raise
    
    def append_expense_data(self, expense_data_list, file_path=None):
        """
        既存のスプレッドシートに経費データを追加
        
        Args:
            expense_data_list (list): 経費データのリスト
            file_path (str, optional): 既存のファイルパス。指定しない場合はインスタンス作成時のパスを使用。
            
        Returns:
            str: 保存されたファイルのパス
        """
        if file_path is not None:
            self.output_path = file_path
        
        try:
            # 既存のワークブックを読み込み
            workbook = self.load_workbook(self.output_path)
            sheet = workbook.active
            
            # 最終行の取得
            last_row = sheet.max_row
            
            # データの追加
            start_row = last_row + 1
            total_amount = 0
            
            # 既存の合計行を削除（存在する場合）
            if last_row > 1:
                for row in range(last_row, 0, -1):
                    if sheet.cell(row=row, column=3).value == "合計":
                        sheet.delete_rows(row)
                        last_row = row - 1
                        break
            
            # 既存のデータの合計金額を計算
            for row in range(2, last_row + 1):
                cell_value = sheet.cell(row=row, column=4).value
                if isinstance(cell_value, (int, float)):
                    total_amount += cell_value
            
            # 新しいデータの追加
            for i, expense in enumerate(expense_data_list, 1):
                row = last_row + i  # 既存データの後から開始
                
                # No.
                sheet.cell(row=row, column=1).value = last_row - 1 + i  # 既存のNo.の続きから
                
                # 日付
                sheet.cell(row=row, column=2).value = expense.get("date", "")
                
                # 説明
                sheet.cell(row=row, column=3).value = expense.get("description", "")
                
                # 金額
                amount = expense.get("amount", 0)
                sheet.cell(row=row, column=4).value = amount
                # 数値フォーマットを適用（通貨形式）
                sheet.cell(row=row, column=4).number_format = "¥#,##0"
                
                # 備考（カテゴリなど）
                sheet.cell(row=row, column=5).value = expense.get("category", "")
                
                # 合計金額の計算
                total_amount += amount
            
            # 新しいデータ行にスタイルを適用
            if expense_data_list:
                self.apply_data_style(sheet, last_row + 1, last_row + len(expense_data_list), 1, 5)
            
            # 合計行の追加
            total_row = last_row + len(expense_data_list) + 2
            sheet.cell(row=total_row, column=3).value = "合計"
            sheet.cell(row=total_row, column=4).value = total_amount
            sheet.cell(row=total_row, column=4).number_format = "¥#,##0"
            
            # 合計行にスタイルを適用
            self.apply_total_style(sheet, total_row, 4)
            
            # 列幅の調整
            self.adjust_column_width(sheet)
            
            # ファイルの保存
            workbook.save(self.output_path)
            logger.info(f"Expense data appended to {self.output_path}")
            
            return self.output_path
        except Exception as e:
            logger.error(f"Error appending expense data: {str(e)}")
            raise

# 単体テスト用のコード
if __name__ == "__main__":
    # テスト用の経費データ
    test_expenses = [
        {
            "date": "2024-03-13",
            "description": "タクシー代",
            "amount": 1500,
            "category": "交通費"
        },
        {
            "date": "2024-03-13",
            "description": "昼食代",
            "amount": 2800,
            "category": "飲食費"
        },
        {
            "date": "2024-03-13",
            "description": "会議室利用料",
            "amount": 15000,
            "category": "会議費"
        }
    ]
    
    # 新規ファイルへの書き込みテスト
    spreadsheet = ExpenseSpreadsheet("test_expense_report.xlsx")
    try:
        output_path = spreadsheet.write_expense_data(test_expenses)
        print(f"経費データを {output_path} に書き込みました")
        
        # 既存ファイルへの追加テスト
        additional_expenses = [
            {
                "date": "2024-03-14",
                "description": "文房具",
                "amount": 3200,
                "category": "消耗品費"
            }
        ]
        
        output_path = spreadsheet.append_expense_data(additional_expenses)
        print(f"追加の経費データを {output_path} に追記しました")
    except Exception as e:
        print(f"エラーが発生しました: {str(e)}")
